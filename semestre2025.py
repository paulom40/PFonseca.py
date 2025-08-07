import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Relatório Interativo", layout="wide")
st.title("📈 Relatório Interativo - KPIs do 1º Semestre 2025")

@st.cache_data(ttl=3600)
def load_data():
    url = "https://github.com/paulom40/PFonseca.py/raw/main/1Semestre2025.xlsx"
    response = requests.get(url)
    df = pd.read_excel(BytesIO(response.content), sheet_name="Dados")
    return df

try:
    df = load_data()
    st.success("✅ Dados carregados com sucesso!")
except Exception as e:
    st.error(f"❌ Erro ao carregar os dados: {e}")
    st.stop()

# Sidebar filters
st.sidebar.header("Filtros")
ano_selecionado = st.sidebar.multiselect("Ano", sorted(df['Ano'].dropna().unique()), default=sorted(df['Ano'].dropna().unique()))
mes_selecionado = st.sidebar.multiselect("Mês", sorted(df['Mês'].dropna().unique()), default=sorted(df['Mês'].dropna().unique()))
artigo_selecionado = st.sidebar.multiselect("Artigo", sorted(df['Artigo'].dropna().unique()), default=sorted(df['Artigo'].dropna().unique()))
comercial_selecionado = st.sidebar.multiselect("Comercial", sorted(df['Comercial'].dropna().unique()), default=sorted(df['Comercial'].dropna().unique()))
cliente_selecionado = st.sidebar.multiselect("Cliente", sorted(df['Cliente'].dropna().unique()), default=sorted(df['Cliente'].dropna().unique()))

# Apply filters
df_filtrado = df[
    (df['Ano'].isin(ano_selecionado)) &
    (df['Mês'].isin(mes_selecionado)) &
    (df['Artigo'].isin(artigo_selecionado)) &
    (df['Comercial'].isin(comercial_selecionado)) &
    (df['Cliente'].isin(cliente_selecionado))
]

# Show filtered data
st.subheader("📊 Tabela de Dados Filtrados")
st.dataframe(df_filtrado, use_container_width=True)

# Summary
st.subheader("📌 Resumo")
st.write(f"Total de Registros: {len(df_filtrado)}")
if 'Valor' in df_filtrado.columns:
    st.write(f"Valor Total: €{df_filtrado['Valor'].sum():,.2f}")

# KPIs by Mês
st.header("📌 KPIs Mensais")

for mes in sorted(df_filtrado['Mês'].dropna().unique()):
    st.subheader(f"📅 Mês: {mes}")
    df_mes = df_filtrado[df_filtrado['Mês'] == mes]

    if 'Artigo' in df_mes.columns and 'Kgs' in df_mes.columns:
        top_artigos = df_mes.groupby('Artigo')['Kgs'].sum().sort_values(ascending=False).head(10)
        st.markdown("**Top 10 Artigos (por Kgs):**")
        st.dataframe(top_artigos.reset_index(), use_container_width=True)

    if 'Cliente' in df_mes.columns and 'Valor' in df_mes.columns:
        top_clientes = df_mes.groupby('Cliente')['Valor'].sum().sort_values(ascending=False).head(10)
        top_clientes_formatted = top_clientes.apply(lambda x: f"€{x:,.2f}")
        st.markdown("**Top 10 Clientes (por Valor):**")
        st.dataframe(top_clientes_formatted.reset_index(), use_container_width=True)

    if 'Comercial' in df_mes.columns and 'Valor' in df_mes.columns:
        top_comerciais = df_mes.groupby('Comercial')['Valor'].sum().sort_values(ascending=False).head(5)
        top_comerciais_formatted = top_comerciais.apply(lambda x: f"€{x:,.2f}")
        st.markdown("**Top 5 Comerciais (por Valor):**")
        st.dataframe(top_comerciais_formatted.reset_index(), use_container_width=True)

# Excel export with charts
def to_excel_with_kpis_and_charts(df_filtrado):
    wb = Workbook()
    ws_filtrado = wb.active
    ws_filtrado.title = "Filtrado"
    for r in dataframe_to_rows(df_filtrado, index=False, header=True):
        ws_filtrado.append(r)

    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Total de Registros", len(df_filtrado)])
    ws_resumo.append(["Valor Total (€)", df_filtrado['Valor'].sum()])

    for mes in sorted(df_filtrado['Mês'].dropna().unique()):
        df_mes = df_filtrado[df_filtrado['Mês'] == mes]

        # Top Artigos
        top_artigos = df_mes.groupby('Artigo')['Kgs'].sum().sort_values(ascending=False).head(10).reset_index()
        ws_artigos = wb.create_sheet(f"{mes}_Artigos")
        for r in dataframe_to_rows(top_artigos, index=False, header=True):
            ws_artigos.append(r)

        chart = BarChart()
        chart.title = f"Top 10 Artigos - {mes}"
        chart.x_axis.title = "Artigo"
        chart.y_axis.title = "Kgs"
        data = Reference(ws_artigos, min_col=2, min_row=1, max_row=11)
        categories = Reference(ws_artigos, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws_artigos.add_chart(chart, "E2")

        # Top Clientes
        top_clientes = df_mes.groupby('Cliente')['Valor'].sum().sort_values(ascending=False).head(10).reset_index()
        ws_clientes = wb.create_sheet(f"{mes}_Clientes")
        for r in dataframe_to_rows(top_clientes, index=False, header=True):
            ws_clientes.append(r)

        chart = BarChart()
        chart.title = f"Top 10 Clientes - {mes}"
        chart.x_axis.title = "Cliente"
        chart.y_axis.title = "Valor (€)"
        data = Reference(ws_clientes, min_col=2, min_row=1, max_row=11)
        categories = Reference(ws_clientes, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws_clientes.add_chart(chart, "E2")

        # Top Comerciais
        top_comerciais = df_mes.groupby('Comercial')['Valor'].sum().sort_values(ascending=False).head(5).reset_index()
        ws_comerciais = wb.create_sheet(f"{mes}_Comerciais")
        for r in dataframe_to_rows(top_comerciais, index=False, header=True):
            ws_comerciais.append(r)

        chart = BarChart()
        chart.title = f"Top 5 Comerciais - {mes}"
        chart.x_axis.title = "Comercial"
        chart.y_axis.title = "Valor (€)"
        data = Reference(ws_comerciais, min_col=2, min_row=1, max_row=6)
        categories = Reference(ws_comerciais, min_col=1, min_row=2, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws_comerciais.add_chart(chart, "E2")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

excel_data = to_excel_with_kpis_and_charts(df_filtrado)

st.download_button(
    label="📥 Baixar Relatório com Gráficos",
    data=excel_data,
    file_name="relatorio_kpis_com_graficos.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
