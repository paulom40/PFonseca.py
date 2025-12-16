import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
import io
from openpyxl import Workbook

# ====================== CONFIG STREAMLIT ======================
st.set_page_config(
    page_title="Dashboard Comercial",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ====================== CSS ======================
st.markdown("""
<style>
    .metric-container {
        background-color: #f5f5f5;
        padding: 15px;
        border-radius: 10px;
        text-align: center;
        border: 1px solid #ddd;
    }
    .metric-value {
        font-size: 26px;
        font-weight: bold;
        color: #333;
    }
    .metric-label {
        font-size: 14px;
        color: #666;
    }
</style>
""", unsafe_allow_html=True)

# ====================== HEADER ======================
st.title("📊 Dashboard Comercial — Vendas & KPIs")
st.markdown("Análise completa de vendas, comerciais, clientes e produtos.")

# ====================== LOAD DATA ======================
@st.cache_data
def load_data(path_or_file="ResumoTR.xlsx") -> pd.DataFrame:
    try:
        df = pd.read_excel(path_or_file)
    except Exception as e:
        st.error(f"Erro a carregar o ficheiro de dados: {e}")
        return pd.DataFrame()

    df.columns = [str(c).strip() for c in df.columns]

    col_map = {
        "Entidad": "Entidade", "Entidade": "Entidade",
        "Nome": "Nome", "Artigo": "Artigo",
        "Cantidad": "Quantidade", "Quantidad": "Quantidade", "Quantidade": "Quantidade",
        "Unidad": "Unidade", "Unidade": "Unidade",
        "V Líquid": "V Líquido", "V_Liquid": "V Líquido", "V Líquido": "V Líquido",
        "PM": "PM", "Data": "Data", "Comercial": "Comercial",
        "Mês": "Mês", "Mes": "Mês", "Ano": "Ano"
    }
    df = df.rename(columns={c: col_map.get(c, c) for c in df.columns})

    required = ["Entidade","Nome","Artigo","Quantidade","Unidade","V Líquido","PM","Data","Comercial","Mês","Ano"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltam colunas obrigatórias: {missing}")
        return pd.DataFrame()

    df["Quantidade"] = (
        df["Quantidade"].astype(str)
        .str.replace(",", ".", regex=False)
        .str.replace("KG", "", regex=False)
        .str.replace("kg", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace("\u00A0", "", regex=False)
    )
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce")
    df["V Líquido"] = pd.to_numeric(df["V Líquido"], errors="coerce")
    df["PM"] = pd.to_numeric(df["PM"], errors="coerce")
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")

    df = df.dropna(subset=["Data"])
    df = df[(df["Quantidade"] > 0) & (df["V Líquido"] != 0)]
    df["AnoMes"] = df["Data"].dt.strftime("%Y-%m")
    return df


# ====================== FILTROS ======================
def aplicar_filtros(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.header("Filtros")
    if df.empty:
        return df

    data_min, data_max = df["Data"].min().date(), df["Data"].max().date()
    data_inicio, data_fim = st.sidebar.date_input("Período", value=(data_min, data_max))

    mask = (df["Data"].dt.date >= data_inicio) & (df["Data"].dt.date <= data_fim)
    df_filt = df[mask].copy()

    # Comercial
    coms = sorted(df_filt["Comercial"].dropna().unique())
    sel_com = st.sidebar.multiselect("Comercial", options=coms, default=coms)
    if sel_com:
        df_filt = df_filt[df_filt["Comercial"].isin(sel_com)]

    # Artigo
    arts = sorted(df_filt["Artigo"].dropna().unique())
    sel_art = st.sidebar.multiselect("Artigo", options=arts, default=arts)
    if sel_art:
        df_filt = df_filt[df_filt["Artigo"].isin(sel_art)]

    # Nome entidade
    df_filt["Nome"] = df_filt["Nome"].astype(str).fillna("").str.strip()
    nomes = sorted([n for n in df_filt["Nome"].unique() if n and n.lower() != "nan"])
    sel_nome = st.sidebar.multiselect("Nome entidade", options=nomes, default=nomes)
    if sel_nome:
        df_filt = df_filt[df_filt["Nome"].isin(sel_nome)]

    return df_filt

# ====================== KPIs ======================
def calcular_kpis(df: pd.DataFrame) -> dict:
    if df.empty:
        return {
            "total_vendas": 0,
            "qtd": 0,
            "clientes": 0,
            "produtos": 0,
            "trans": 0,
            "ticket_medio": 0,
            "ticket_medio_cliente": 0,
            "venda_dia": 0,
            "valor_unidade": 0,
            "periodo": "Sem dados"
        }

    total_vendas = df["V Líquido"].sum()
    qtd_total = df["Quantidade"].sum()
    periodo = f"{df['Data'].min().strftime('%d/%m/%Y')} a {df['Data'].max().strftime('%d/%m/%Y')}"
    dias_com_venda = df["Data"].dt.date.nunique()

    # CORREÇÃO: Transações = Agrupamento por Cliente + Data (visitas únicas)
    df_temp = df.copy()
    df_temp["Data_Apenas"] = df_temp["Data"].dt.date
    trans = df_temp.groupby(["Nome", "Data_Apenas"]).ngroups  # Cada cliente+dia = 1 transação
    
    clientes = df["Nome"].nunique()
    produtos = df["Artigo"].nunique()

    # Ticket Médio = Valor médio por transação (visita)
    ticket_medio = total_vendas / trans if trans else 0
    
    # Ticket Médio por Cliente = Valor médio de compra por cliente
    ticket_medio_cliente = total_vendas / clientes if clientes else 0
    
    # Venda média por dia
    venda_dia = total_vendas / dias_com_venda if dias_com_venda else 0
    
    # Valor médio por unidade vendida
    valor_unidade = total_vendas / qtd_total if qtd_total else 0

    return {
        "total_vendas": total_vendas,
        "qtd": qtd_total,
        "clientes": clientes,
        "produtos": produtos,
        "trans": trans,
        "ticket_medio": ticket_medio,
        "ticket_medio_cliente": ticket_medio_cliente,
        "venda_dia": venda_dia,
        "valor_unidade": valor_unidade,
        "periodo": periodo
    }


# ====================== ANÁLISE DE CRESCIMENTO ======================
def calcular_crescimento(df: pd.DataFrame) -> dict:
    """
    Calcula crescimento comparando último mês com mês anterior
    """
    if df.empty:
        return None
    
    df = df.copy()
    df["AnoMes"] = df["Data"].dt.strftime("%Y-%m")
    df["Data_Apenas"] = df["Data"].dt.date
    
    meses = sorted(df["AnoMes"].unique())
    if len(meses) < 2:
        return None
    
    # Último mês e mês anterior
    mes_atual = meses[-1]
    mes_anterior = meses[-2]
    
    df_atual = df[df["AnoMes"] == mes_atual]
    df_anterior = df[df["AnoMes"] == mes_anterior]
    
    # Calcular KPIs de ambos períodos
    vendas_atual = df_atual["V Líquido"].sum()
    vendas_anterior = df_anterior["V Líquido"].sum()
    
    qtd_atual = df_atual["Quantidade"].sum()
    qtd_anterior = df_anterior["Quantidade"].sum()
    
    # CORREÇÃO: Contar transações como visitas (Cliente + Data)
    trans_atual = df_atual.groupby(["Nome", "Data_Apenas"]).ngroups
    trans_anterior = df_anterior.groupby(["Nome", "Data_Apenas"]).ngroups
    
    clientes_atual = df_atual["Nome"].nunique()
    clientes_anterior = df_anterior["Nome"].nunique()
    
    # Calcular variações %
    var_vendas = ((vendas_atual - vendas_anterior) / vendas_anterior * 100) if vendas_anterior else 0
    var_qtd = ((qtd_atual - qtd_anterior) / qtd_anterior * 100) if qtd_anterior else 0
    var_trans = ((trans_atual - trans_anterior) / trans_anterior * 100) if trans_anterior else 0
    var_clientes = ((clientes_atual - clientes_anterior) / clientes_anterior * 100) if clientes_anterior else 0
    
    return {
        "mes_atual": mes_atual,
        "mes_anterior": mes_anterior,
        "vendas_atual": vendas_atual,
        "vendas_anterior": vendas_anterior,
        "var_vendas": var_vendas,
        "qtd_atual": qtd_atual,
        "qtd_anterior": qtd_anterior,
        "var_qtd": var_qtd,
        "trans_atual": trans_atual,
        "trans_anterior": trans_anterior,
        "var_trans": var_trans,
        "clientes_atual": clientes_atual,
        "clientes_anterior": clientes_anterior,
        "var_clientes": var_clientes
    }


def mostrar_analise_crescimento(df: pd.DataFrame):
    st.subheader("📈 Análise de Crescimento (Mês vs Mês Anterior)")
    
    crescimento = calcular_crescimento(df)
    
    if not crescimento:
        st.warning("Necessário pelo menos 2 meses de dados para análise de crescimento.")
        return
    
    st.info(f"Comparando **{crescimento['mes_atual']}** vs **{crescimento['mes_anterior']}**")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Total Vendas",
            f"{crescimento['vendas_atual']:,.2f}€",
            f"{crescimento['var_vendas']:+.1f}%",
            delta_color="normal"
        )
    
    with col2:
        st.metric(
            "Quantidade",
            f"{crescimento['qtd_atual']:,.0f}",
            f"{crescimento['var_qtd']:+.1f}%",
            delta_color="normal"
        )
    
    with col3:
        st.metric(
            "Transações",
            f"{crescimento['trans_atual']:,}",
            f"{crescimento['var_trans']:+.1f}%",
            delta_color="normal"
        )
    
    with col4:
        st.metric(
            "Clientes",
            f"{crescimento['clientes_atual']:,}",
            f"{crescimento['var_clientes']:+.1f}%",
            delta_color="normal"
        )
    
    # Gráfico de comparação
    df_comp = pd.DataFrame({
        "Período": [crescimento['mes_anterior'], crescimento['mes_atual']],
        "Vendas": [crescimento['vendas_anterior'], crescimento['vendas_atual']],
        "Transações": [crescimento['trans_anterior'], crescimento['trans_atual']],
        "Clientes": [crescimento['clientes_anterior'], crescimento['clientes_atual']]
    })
    
    fig = px.bar(
        df_comp,
        x="Período",
        y="Vendas",
        text="Vendas",
        title="Comparação de Vendas",
        color="Vendas",
        color_continuous_scale="Blues"
    )
    fig.update_traces(texttemplate="%{text:,.0f}€", textposition="outside")
    fig.update_layout(height=400)
    st.plotly_chart(fig, width='stretch')


# ====================== SEGMENTAÇÃO DE CLIENTES ======================
def segmentar_clientes(df: pd.DataFrame) -> pd.DataFrame:
    """
    Segmenta clientes em: VIP, Regulares, Novos, Inativos
    
    Critérios:
    - VIP: Top 20% em vendas E mais de 5 transações (visitas)
    - Regulares: Mais de 3 transações (visitas)
    - Novos: 1-3 transações E primeira compra nos últimos 90 dias
    - Inativos: Última compra há mais de 90 dias
    """
    if df.empty:
        return pd.DataFrame()
    
    df = df.copy()
    df["Data_Apenas"] = df["Data"].dt.date
    data_max = df["Data"].max()
    
    # Agregar por cliente
    clientes = df.groupby("Nome").agg(
        Total_Vendas=("V Líquido", "sum"),
        Primeira_Compra=("Data", "min"),
        Ultima_Compra=("Data", "max"),
        Quantidade=("Quantidade", "sum")
    ).reset_index()
    
    # CORREÇÃO: Contar transações como visitas únicas (Cliente + Data)
    transacoes_por_cliente = df.groupby("Nome").apply(
        lambda x: x.groupby("Data_Apenas").ngroups
    ).reset_index(name="Transacoes")
    
    clientes = clientes.merge(transacoes_por_cliente, on="Nome")
    
    # Calcular dias desde última compra
    clientes["Dias_Sem_Comprar"] = (data_max - clientes["Ultima_Compra"]).dt.days
    clientes["Dias_Como_Cliente"] = (data_max - clientes["Primeira_Compra"]).dt.days
    
    # Threshold para top 20%
    percentil_80 = clientes["Total_Vendas"].quantile(0.80)
    
    # Função de segmentação
    def classificar(row):
        if row["Dias_Sem_Comprar"] > 90:
            return "⚠️ Inativo"
        elif row["Total_Vendas"] >= percentil_80 and row["Transacoes"] > 5:
            return "⭐ VIP"
        elif row["Transacoes"] > 3:
            return "✅ Regular"
        elif row["Dias_Como_Cliente"] <= 90:
            return "🆕 Novo"
        else:
            return "📊 Ocasional"
    
    clientes["Segmento"] = clientes.apply(classificar, axis=1)
    
    # Calcular ticket médio
    clientes["Ticket_Medio"] = clientes["Total_Vendas"] / clientes["Transacoes"]
    
    return clientes


def mostrar_segmentacao_clientes(df: pd.DataFrame):
    st.subheader("👥 Segmentação de Clientes")
    
    if df.empty:
        st.warning("Sem dados para segmentação.")
        return
    
    clientes_seg = segmentar_clientes(df)
    
    if clientes_seg.empty:
        st.warning("Não foi possível segmentar clientes.")
        return
    
    # Resumo por segmento
    resumo = clientes_seg.groupby("Segmento").agg(
        Quantidade_Clientes=("Nome", "count"),
        Total_Vendas=("Total_Vendas", "sum"),
        Ticket_Medio=("Ticket_Medio", "mean")
    ).reset_index()
    
    # Ordenar por importância
    ordem_segmentos = ["⭐ VIP", "✅ Regular", "🆕 Novo", "📊 Ocasional", "⚠️ Inativo"]
    resumo["Ordem"] = resumo["Segmento"].map({s: i for i, s in enumerate(ordem_segmentos)})
    resumo = resumo.sort_values("Ordem").drop("Ordem", axis=1)
    
    # Gráfico de pizza
    col1, col2 = st.columns(2)
    
    with col1:
        fig_pizza = px.pie(
            resumo,
            values="Quantidade_Clientes",
            names="Segmento",
            title="Distribuição de Clientes por Segmento",
            hole=0.4
        )
        fig_pizza.update_layout(height=400)
        st.plotly_chart(fig_pizza, width='stretch')
    
    with col2:
        fig_vendas = px.bar(
            resumo,
            x="Segmento",
            y="Total_Vendas",
            text="Total_Vendas",
            title="Vendas por Segmento",
            color="Total_Vendas",
            color_continuous_scale="Viridis"
        )
        fig_vendas.update_traces(texttemplate="%{text:,.0f}€", textposition="outside")
        fig_vendas.update_layout(height=400)
        st.plotly_chart(fig_vendas, width='stretch')
    
    # Tabela resumo
    st.markdown("### 📊 Resumo por Segmento")
    resumo_display = resumo.copy()
    resumo_display["Total_Vendas"] = resumo_display["Total_Vendas"].map(lambda x: f"{x:,.2f}€")
    resumo_display["Ticket_Medio"] = resumo_display["Ticket_Medio"].map(lambda x: f"{x:,.2f}€")
    resumo_display.columns = ["Segmento", "Nº Clientes", "Total Vendas", "Ticket Médio"]
    st.dataframe(resumo_display, width='stretch', hide_index=True)
    
    # Lista de clientes VIP
    st.markdown("### ⭐ Clientes VIP")
    vips = clientes_seg[clientes_seg["Segmento"] == "⭐ VIP"].sort_values("Total_Vendas", ascending=False)
    
    if not vips.empty:
        vips_display = vips[["Nome", "Total_Vendas", "Transacoes", "Ticket_Medio", "Ultima_Compra"]].copy()
        vips_display["Total_Vendas"] = vips_display["Total_Vendas"].map(lambda x: f"{x:,.2f}€")
        vips_display["Ticket_Medio"] = vips_display["Ticket_Medio"].map(lambda x: f"{x:,.2f}€")
        vips_display["Ultima_Compra"] = vips_display["Ultima_Compra"].dt.strftime("%d/%m/%Y")
        vips_display.columns = ["Cliente", "Total Vendas", "Transações", "Ticket Médio", "Última Compra"]
        st.dataframe(vips_display, width='stretch', hide_index=True)
    else:
        st.info("Nenhum cliente VIP no período selecionado.")
    
    # Clientes inativos (em risco)
    st.markdown("### ⚠️ Clientes Inativos (Recuperar)")
    inativos = clientes_seg[clientes_seg["Segmento"] == "⚠️ Inativo"].sort_values("Total_Vendas", ascending=False)
    
    if not inativos.empty:
        inativos_display = inativos[["Nome", "Total_Vendas", "Dias_Sem_Comprar", "Ultima_Compra"]].head(10).copy()
        inativos_display["Total_Vendas"] = inativos_display["Total_Vendas"].map(lambda x: f"{x:,.2f}€")
        inativos_display["Ultima_Compra"] = inativos_display["Ultima_Compra"].dt.strftime("%d/%m/%Y")
        inativos_display.columns = ["Cliente", "Total Vendas (histórico)", "Dias sem comprar", "Última Compra"]
        st.dataframe(inativos_display, width='stretch', hide_index=True)
    else:
        st.success("Nenhum cliente inativo! 🎉")
    
    # Botão de download
    st.markdown("---")
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        clientes_export = clientes_seg.copy()
        clientes_export["Primeira_Compra"] = clientes_export["Primeira_Compra"].dt.strftime("%d/%m/%Y")
        clientes_export["Ultima_Compra"] = clientes_export["Ultima_Compra"].dt.strftime("%d/%m/%Y")
        clientes_export = clientes_export[[
            "Nome", "Segmento", "Total_Vendas", "Transacoes", 
            "Ticket_Medio", "Dias_Sem_Comprar", "Primeira_Compra", "Ultima_Compra"
        ]]
        clientes_export.columns = [
            "Cliente", "Segmento", "Total Vendas (€)", "Transações", 
            "Ticket Médio (€)", "Dias sem Comprar", "Primeira Compra", "Última Compra"
        ]
        clientes_export.to_excel(writer, sheet_name="Segmentacao_Clientes", index=False)
    
    buffer.seek(0)
    st.download_button(
        label="📥 Download Segmentação de Clientes (Excel)",
        data=buffer,
        file_name="Segmentacao_Clientes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
def calcular_ticket_medio_por_comercial(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    grp = df.groupby("Comercial").agg(
        Total_Vendas=("V Líquido", "sum"),
        Transacoes=("V Líquido", "count"),
        Quantidade=("Quantidade", "sum")
    ).reset_index()

    grp["Ticket_Medio"] = grp["Total_Vendas"] / grp["Transacoes"]
    grp["Valor_Medio_Unidade"] = grp["Total_Vendas"] / grp["Quantidade"]

    return grp.sort_values("Total_Vendas", ascending=False)

# ====================== VISUALIZAÇÕES PRINCIPAIS ======================
def desenhar_kpis(kpis: dict, df_ticket_com: pd.DataFrame):
    st.subheader("KPIs em Tempo Real")
    
    # Debug info (pode remover depois)
    with st.expander("ℹ️ Informação de Debug"):
        st.write(f"**Total Vendas:** {kpis['total_vendas']:,.2f}€")
        st.write(f"**Número de Transações (linhas):** {kpis['trans']}")
        st.write(f"**Ticket Médio Calculado:** {kpis['total_vendas'] / kpis['trans'] if kpis['trans'] else 0:,.2f}€")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Vendas (€)", f"{kpis['total_vendas']:,.2f}")
    c2.metric("Quantidade Total", f"{kpis['qtd']:,.2f}")
    c3.metric("Clientes Únicos", int(kpis["clientes"]))
    c4.metric("Produtos Vendidos", int(kpis["produtos"]))
    c5.metric("Transações", int(kpis["trans"]))

    st.divider()

    c6, c7, c8 = st.columns(3)
    c6.metric("Ticket Médio (€)", f"{kpis['ticket_medio']:,.2f}")
    c7.metric("Ticket Médio por Cliente (€)", f"{kpis['ticket_medio_cliente']:,.2f}")
    c8.metric("Valor Médio por Unidade (€)", f"{kpis['valor_unidade']:,.4f}")

    st.info(f"Período em análise: {kpis['periodo']}")

    st.subheader("📊 Desempenho por Comercial")
    if df_ticket_com.empty:
        st.warning("Sem dados de comercial.")
    else:
        # Verificar se a coluna Transacoes existe
        if "Transacoes" not in df_ticket_com.columns:
            st.error("Erro: Coluna 'Transacoes' não encontrada. Verifique a função calcular_ticket_medio_por_comercial()")
            st.write("Colunas disponíveis:", df_ticket_com.columns.tolist())
            return
        
        df_show = df_ticket_com.copy()
        
        # Debug: Mostrar valores brutos
        with st.expander("🔍 Debug - Valores Brutos"):
            st.dataframe(df_show)
        
        # Formatar colunas para display
        df_show["Total_Vendas_Format"] = df_show["Total_Vendas"].map(lambda x: f"{x:,.2f}€")
        df_show["Transacoes_Format"] = df_show["Transacoes"].map(lambda x: f"{int(x)}")
        df_show["Quantidade_Format"] = df_show["Quantidade"].map(lambda x: f"{x:,.3f}")
        df_show["Ticket_Medio_Format"] = df_show["Ticket_Medio"].map(lambda x: f"{x:,.2f}€")
        df_show["Valor_Medio_Unidade_Format"] = df_show["Valor_Medio_Unidade"].map(lambda x: f"{x:,.4f}€")
        
        # Selecionar colunas formatadas
        df_display = df_show[["Comercial", "Total_Vendas_Format", "Transacoes_Format", 
                              "Quantidade_Format", "Ticket_Medio_Format", "Valor_Medio_Unidade_Format"]]
        df_display.columns = ["Comercial", "Total Vendas", "Transações", 
                             "Quantidade", "Ticket Médio", "Valor Médio/Unidade"]
        
        st.dataframe(df_display, width='stretch', hide_index=True)
        
        # Botão de download
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Preparar dados para Excel (com valores numéricos originais)
            df_excel = df_show[["Comercial", "Total_Vendas", "Transacoes", 
                                "Quantidade", "Ticket_Medio", "Valor_Medio_Unidade"]].copy()
            df_excel.columns = ["Comercial", "Total Vendas (€)", "Transações", 
                               "Quantidade", "Ticket Médio (€)", "Valor Médio/Unidade (€)"]
            df_excel.to_excel(writer, sheet_name="Desempenho_Comerciais", index=False)
        
        buffer.seek(0)
        
        st.download_button(
            label="📥 Download Desempenho por Comercial (Excel)",
            data=buffer,
            file_name="Desempenho_Comerciais.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Mostrar totais
        st.markdown("---")
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Total Geral", f"{df_show['Total_Vendas'].sum():,.2f}€")
        col_b.metric("Total Transações", f"{int(df_show['Transacoes'].sum())}")
        col_c.metric("Total Quantidade", f"{df_show['Quantidade'].sum():,.3f}")


def grafico_evolucao(df: pd.DataFrame):
    st.subheader("Evolução Mensal de Vendas (€)")
    if df.empty:
        st.warning("Sem dados.")
        return

    mensal = df.groupby("AnoMes")["V Líquido"].sum().reset_index()

    fig = px.line(mensal, x="AnoMes", y="V Líquido", markers=True)
    fig.add_bar(x=mensal["AnoMes"], y=mensal["V Líquido"])
    fig.update_layout(height=500)

    st.plotly_chart(fig, width='stretch')


def graficos_top10(df: pd.DataFrame):
    col1, col2 = st.columns(2)

    # Top 10 Produtos (€)
    with col1:
        st.subheader("Top 10 Produtos (€)")
        if df.empty:
            st.warning("Sem dados.")
        else:
            topp = df.groupby("Artigo")["V Líquido"].sum().nlargest(10)
            figp = px.bar(
                x=topp.values, y=topp.index, orientation="h",
                color=topp.values, color_continuous_scale="Plasma"
            )
            figp.update_layout(height=500)
            st.plotly_chart(figp, width='stretch')

    # Top 10 Clientes (€)
    with col2:
        st.subheader("Top 10 Clientes (€)")
        if df.empty:
            st.warning("Sem dados.")
        else:
            topc = df.groupby("Nome")["V Líquido"].sum().nlargest(10)
            figc = px.bar(
                x=topc.values, y=topc.index, orientation="h",
                color=topc.values, color_continuous_scale="Viridis"
            )
            figc.update_layout(height=500)
            st.plotly_chart(figc, width='stretch')

    st.divider()

    col3, col4 = st.columns(2)

    # Top 10 Produtos (Quantidade)
    with col3:
        st.subheader("Top 10 Produtos (Quantidade)")
        if df.empty:
            st.warning("Sem dados.")
        else:
            topp_q = df.groupby("Artigo")["Quantidade"].sum().nlargest(10)
            figpq = px.bar(
                x=topp_q.values, y=topp_q.index, orientation="h",
                color=topp_q.values, color_continuous_scale="Blues"
            )
            figpq.update_layout(height=500)
            st.plotly_chart(figpq, width='stretch')

    # Top 10 Clientes (Quantidade)
    with col4:
        st.subheader("Top 10 Clientes (Quantidade)")
        if df.empty:
            st.warning("Sem dados.")
        else:
            topc_q = df.groupby("Nome")["Quantidade"].sum().nlargest(10)
            figcq = px.bar(
                x=topc_q.values, y=topc_q.index, orientation="h",
                color=topc_q.values, color_continuous_scale="Greens"
            )
            figcq.update_layout(height=500)
            st.plotly_chart(figcq, width='stretch')

# ====================== COMPARAÇÃO MÊS A MÊS (GLOBAL) ======================
def comparacao_mes_a_mes(df: pd.DataFrame):
    st.subheader("📆 Comparação Mês a Mês — Global")

    if df.empty:
        st.warning("Sem dados para comparar.")
        return

    df = df.copy()
    df["Ano"] = df["Data"].dt.year

    anos_disponiveis = sorted(df["Ano"].unique())
    if len(anos_disponiveis) < 1:
        st.warning("Sem anos disponíveis.")
        return

    ano_sel = st.selectbox(
        "Seleciona o ano:",
        options=anos_disponiveis,
        key="ano_mes_a_mes_global"
    )

    df_ano = df[df["Ano"] == ano_sel]

    df_comp = df_ano.groupby("AnoMes").agg(
        Total_Vendas=("V Líquido", "sum"),
        Quantidade=("Quantidade", "sum"),
        Transacoes=("V Líquido", "count"),
        Clientes=("Nome", "nunique")
    ).reset_index()

    # Ordenar por mês
    df_comp = df_comp.sort_values("AnoMes")

    fig = px.bar(
        df_comp,
        x="AnoMes",
        y="Total_Vendas",
        text="Total_Vendas",
        title=f"Vendas Mês a Mês — {ano_sel}",
        color="Total_Vendas",
        color_continuous_scale="Blues"
    )
    fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
    fig.update_layout(height=500, xaxis_title="Mês", yaxis_title="Total Vendas (€)")

    st.plotly_chart(fig, width='stretch')

    # Gráfico de linha para evolução
    fig_line = px.line(
        df_comp,
        x="AnoMes",
        y="Total_Vendas",
        markers=True,
        title=f"Evolução Mensal — {ano_sel}"
    )
    fig_line.update_layout(height=400, xaxis_title="Mês", yaxis_title="Total Vendas (€)")
    st.plotly_chart(fig_line, width='stretch')

    # Tabela de dados
    df_show = df_comp.copy()
    df_show["Total_Vendas"] = df_show["Total_Vendas"].map(lambda x: f"{x:,.2f}")
    df_show["Quantidade"] = df_show["Quantidade"].map(lambda x: f"{x:,.2f}")
    df_show.columns = ["Mês", "Total Vendas (€)", "Quantidade", "Transações", "Clientes"]

    st.dataframe(df_show, width='stretch', hide_index=True)


# ====================== COMPARAÇÃO ANO-A-ANO POR CLIENTE ======================
def comparacao_ano_a_ano_clientes(df: pd.DataFrame):
    st.subheader("👥 Comparação Ano-a-Ano — Clientes")

    if df.empty:
        st.warning("Sem dados.")
        return

    df = df.copy()
    df["Mes_Num"] = df["Data"].dt.month
    df["Ano"] = df["Data"].dt.year

    clientes = sorted(df["Nome"].dropna().unique())
    if not clientes:
        st.warning("Sem clientes disponíveis.")
        return

    cliente_sel = st.selectbox(
        "Seleciona o cliente:",
        options=clientes,
        key="cliente_ano_ano"
    )

    df_cliente = df[df["Nome"] == cliente_sel]

    meses_disponiveis = sorted(df_cliente["Mes_Num"].unique())
    if not meses_disponiveis:
        st.warning("Sem meses disponíveis para este cliente.")
        return

    mes_sel = st.selectbox(
        "Seleciona o mês:",
        options=meses_disponiveis,
        format_func=lambda m: datetime(2000, m, 1).strftime("%B"),
        key="mes_ano_ano_cliente"
    )

    df_mes = df_cliente[df_cliente["Mes_Num"] == mes_sel]

    df_comp = df_mes.groupby("Ano").agg(
        Total_Vendas=("V Líquido", "sum"),
        Quantidade=("Quantidade", "sum"),
        Transacoes=("V Líquido", "count")
    ).reset_index()

    fig = px.bar(
        df_comp,
        x="Ano",
        y="Total_Vendas",
        text="Total_Vendas",
        title=f"{cliente_sel} — {datetime(2000, mes_sel, 1).strftime('%B')} (Ano-a-Ano)",
        color="Total_Vendas",
        color_continuous_scale="Blues"
    )
    fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
    fig.update_layout(height=500)

    st.plotly_chart(fig, width='stretch')

    df_show = df_comp.copy()
    df_show["Total_Vendas"] = df_show["Total_Vendas"].map(lambda x: f"{x:,.2f}")
    df_show["Quantidade"] = df_show["Quantidade"].map(lambda x: f"{x:,.2f}")

    st.dataframe(df_show, width='stretch', hide_index=True)

# ====================== AUXILIARES PARA EXPORTAÇÃO EXCEL ======================

def sanitize_sheet_name(name: str, existing_names: set) -> str:
    """
    Garante que o nome da folha é válido no Excel e único.
    """
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    for ch in invalid_chars:
        name = name.replace(ch, '')

    name = name.strip()
    if len(name) > 31:
        name = name[:31]

    base = name
    counter = 1
    while name in existing_names:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        counter += 1

    existing_names.add(name)
    return name


def criar_sheet(wb, name: str, existing_names: set):
    """
    Cria uma folha com nome validado e único.
    """
    name_real = sanitize_sheet_name(name, existing_names)
    ws = wb.create_sheet(title=name_real)
    return ws

# ====================== EXPORTAÇÃO EXCEL — RELATÓRIO COMPLETO (SEM GRÁFICOS) ======================
def tabela_dados_export(df: pd.DataFrame, kpis: dict):
    st.subheader("📄 Exportação — Relatório Completo")

    if df.empty:
        st.warning("Sem dados para exportar.")
        return

    # Criar workbook
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumo"
    existing_names = {ws0.title}

    # ====================== Folha Resumo ======================
    ws0["A1"] = "Resumo Geral"
    ws0["A3"] = "Total Vendas (€)"
    ws0["B3"] = kpis["total_vendas"]
    ws0["A4"] = "Quantidade Total"
    ws0["B4"] = kpis["qtd"]
    ws0["A5"] = "Clientes Únicos"
    ws0["B5"] = kpis["clientes"]
    ws0["A6"] = "Produtos Vendidos"
    ws0["B6"] = kpis["produtos"]
    ws0["A7"] = "Transações"
    ws0["B7"] = kpis["trans"]
    ws0["A8"] = "Ticket Médio (€)"
    ws0["B8"] = kpis["ticket_medio"]
    ws0["A9"] = "Ticket Médio por Cliente (€)"
    ws0["B9"] = kpis["ticket_medio_cliente"]
    ws0["A10"] = "Venda Média por Dia (€)"
    ws0["B10"] = kpis["venda_dia"]
    ws0["A11"] = "Valor Médio por Unidade (€)"
    ws0["B11"] = kpis["valor_unidade"]

    # ====================== Folha Dados ======================
    ws_dados = criar_sheet(wb, "Dados", existing_names)
    for col_num, col_name in enumerate(df.columns, 1):
        ws_dados.cell(row=1, column=col_num, value=col_name)

    for row_num, row in enumerate(df.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            ws_dados.cell(row=row_num, column=col_num, value=value)

    # ====================== Folha Evolução Mensal (DADOS) ======================
    ws_hist = criar_sheet(wb, "Historico_Mensal", existing_names)
    mensal = df.groupby("AnoMes")["V Líquido"].sum().reset_index()
    
    ws_hist["A1"] = "Mês"
    ws_hist["B1"] = "Vendas (€)"
    for row_num, row in enumerate(mensal.itertuples(index=False), 2):
        ws_hist.cell(row=row_num, column=1, value=row.AnoMes)
        ws_hist.cell(row=row_num, column=2, value=row[1])

    # ====================== Folha Ranking Comerciais (DADOS) ======================
    ws_rank = criar_sheet(wb, "Ranking_Comerciais", existing_names)
    rank = df.groupby("Comercial")["V Líquido"].sum().sort_values(ascending=False).reset_index()
    
    ws_rank["A1"] = "Comercial"
    ws_rank["B1"] = "Total Vendas (€)"
    for row_num, row in enumerate(rank.itertuples(index=False), 2):
        ws_rank.cell(row=row_num, column=1, value=row.Comercial)
        ws_rank.cell(row=row_num, column=2, value=row[1])

    # ====================== Folha Produtos (TOP 10) ======================
    ws_prod = criar_sheet(wb, "Top_Produtos", existing_names)
    top_prod = df.groupby("Artigo")["V Líquido"].sum().nlargest(10).reset_index()
    
    ws_prod["A1"] = "Produto"
    ws_prod["B1"] = "Total Vendas (€)"
    for row_num, row in enumerate(top_prod.itertuples(index=False), 2):
        ws_prod.cell(row=row_num, column=1, value=row.Artigo)
        ws_prod.cell(row=row_num, column=2, value=row[1])

    # ====================== Folha Clientes (TOP 10) ======================
    ws_cli = criar_sheet(wb, "Top_Clientes", existing_names)
    top_cli = df.groupby("Nome")["V Líquido"].sum().nlargest(10).reset_index()
    
    ws_cli["A1"] = "Cliente"
    ws_cli["B1"] = "Total Vendas (€)"
    for row_num, row in enumerate(top_cli.itertuples(index=False), 2):
        ws_cli.cell(row=row_num, column=1, value=row.Nome)
        ws_cli.cell(row=row_num, column=2, value=row[1])

    # ====================== Exportação ======================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="📥 Download Relatório Completo (Excel)",
        data=buffer,
        file_name="Relatorio_Completo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ====================== EXPORTAÇÃO EXCEL — RELATÓRIOS MENSAIS (SEM GRÁFICOS) ======================
def gerar_excel_completo(df: pd.DataFrame):
    st.subheader("📄 Exportação — Relatórios Mensais")

    if df.empty:
        st.warning("Sem dados para exportar.")
        return

    meses = sorted(df["AnoMes"].unique())
    mes_sel = st.selectbox("Seleciona o mês para exportar:", meses)

    df_mes = df[df["AnoMes"] == mes_sel]
    if df_mes.empty:
        st.warning("Sem dados para este mês.")
        return

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumo"
    existing_names = {ws0.title}

    # ====================== Resumo ======================
    ws0["A1"] = f"Resumo — {mes_sel}"
    ws0["A3"] = "Total Vendas (€)"
    ws0["B3"] = df_mes["V Líquido"].sum()
    ws0["A4"] = "Quantidade Total"
    ws0["B4"] = df_mes["Quantidade"].sum()
    ws0["A5"] = "Clientes Únicos"
    ws0["B5"] = df_mes["Nome"].nunique()
    ws0["A6"] = "Produtos Vendidos"
    ws0["B6"] = df_mes["Artigo"].nunique()

    # ====================== Dados ======================
    ws_dados = criar_sheet(wb, "Dados", existing_names)
    for col_num, col_name in enumerate(df_mes.columns, 1):
        ws_dados.cell(row=1, column=col_num, value=col_name)

    for row_num, row in enumerate(df_mes.itertuples(index=False), 2):
        for col_num, value in enumerate(row, 1):
            ws_dados.cell(row=row_num, column=col_num, value=value)

    # ====================== Histórico Diário (DADOS) ======================
    ws_hist = criar_sheet(wb, "Historico_Diario", existing_names)
    diario = df_mes.groupby(df_mes["Data"].dt.strftime("%d"))["V Líquido"].sum().reset_index()
    diario.columns = ["Dia", "Vendas"]

    ws_hist["A1"] = "Dia"
    ws_hist["B1"] = "Vendas (€)"
    for row_num, row in enumerate(diario.itertuples(index=False), 2):
        ws_hist.cell(row=row_num, column=1, value=row.Dia)
        ws_hist.cell(row=row_num, column=2, value=row.Vendas)

    # ====================== Exportação ======================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label=f"📥 Download Relatório {mes_sel}",
        data=buffer,
        file_name=f"Relatorio_{mes_sel}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ====================== INTERFACE DAS EXPORTAÇÕES ======================
def tabelas_export_interface(df: pd.DataFrame, kpis: dict):
    st.subheader("📄 Tabelas & Export")

    st.markdown("### 📥 Exportação Completa")
    tabela_dados_export(df, kpis)

    st.markdown("---")
    st.markdown("### 📥 Exportação Mensal")
    gerar_excel_completo(df)

# ====================== ABA PRINCIPAL — COMPARAÇÕES ======================
def aba_comparacao_ano_ano(df: pd.DataFrame):
    st.header("📆 Comparações Temporais")

    tab1, tab2 = st.tabs([
        "📆 Mês a Mês",
        "👥 Ano-a-Ano (Clientes)"
    ])

    with tab1:
        comparacao_mes_a_mes(df)

    with tab2:
        comparacao_ano_a_ano_clientes(df)

# ====================== ABA PRINCIPAL — DASHBOARD COMPLETO ======================
def aba_dashboard(df: pd.DataFrame):
    st.header("📊 Dashboard Geral")

    if df.empty:
        st.warning("Sem dados para apresentar.")
        return

    # ====================== KPIs ======================
    kpis = calcular_kpis(df)
    df_ticket_com = calcular_ticket_medio_por_comercial(df)

    desenhar_kpis(kpis, df_ticket_com)

    st.divider()

    # ====================== Análise de Crescimento ======================
    mostrar_analise_crescimento(df)

    st.divider()

    # ====================== Segmentação de Clientes ======================
    mostrar_segmentacao_clientes(df)

    st.divider()

    # ====================== Evolução Mensal ======================
    grafico_evolucao(df)

    st.divider()

    # ====================== Top 10 ======================
    graficos_top10(df)

    st.divider()

    # ====================== Tabelas & Export ======================
    tabelas_export_interface(df, kpis)

# ====================== MAIN APP ======================
def main():
    st.sidebar.title("📁 Carregar Dados")

    uploaded_file = st.sidebar.file_uploader(
        "Seleciona o ficheiro Excel",
        type=["xlsx"]
    )

    if uploaded_file:
        df = load_data(uploaded_file)
    else:
        st.info("A usar o ficheiro padrão: ResumoTR.xlsx")
        df = load_data()

    if df.empty:
        st.error("Não foi possível carregar dados válidos.")
        return

    # Aplicar filtros
    df_filt = aplicar_filtros(df)

    # Tabs principais
    tab_dashboard, tab_comparacoes = st.tabs([
        "📊 Dashboard Geral",
        "📆 Comparações Temporais"
    ])

    with tab_dashboard:
        aba_dashboard(df_filt)

    with tab_comparacoes:
        aba_comparacao_ano_ano(df_filt)

# ====================== EXECUÇÃO ======================
if __name__ == "__main__":
    main()
